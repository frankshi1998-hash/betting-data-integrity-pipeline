from __future__ import annotations

import argparse
import json
from datetime import date, datetime, time
from pathlib import Path
from typing import Iterable, Iterator

from src.config import RAW_DATA_DIR, SCHEMA_PATH, get_database_config

HEADER_TO_COLUMN = {
    "REFID": "refid",
    "UUID": "source_uuid",
    "License": "license",
    "Bookmaker": "bookmaker",
    "Location": "location",
    "State": "state",
    "Area": "area",
    "Wagering Provider": "wagering_provider",
    "SysVerNo": "sys_ver_no",
    "Bet Date": "bet_date",
    "Bet Time": "bet_time",
    "Event Date": "event_date",
    "Event Time": "event_time",
    "Ticket No": "ticket_no",
    "Event": "event_category",
    "Sport Event": "sport_event",
    "Venue": "venue",
    "Venue State": "venue_state",
    "Bet Type": "bet_type",
    "Bet Method": "bet_method",
    "Bet Details": "bet_details",
    "Race Number": "race_number",
    "Runner Number": "runner_number",
    "Runner Name": "runner_name",
    "Bet Amount Win": "bet_amount_win",
    "Bet Amount Place": "bet_amount_place",
    "WinPrice": "win_price",
    "Place Price": "place_price",
    "Customer Name": "customer_name",
    "Betback Claim": "betback_claim",
    "Bet Information": "bet_information",
    "Cancelled Flag": "cancelled_flag",
    "Time Cancelled": "time_cancelled",
    "Bet Win Takeout": "bet_win_takeout",
    "Bet Place Takeout": "bet_place_takeout",
    "Horse Win Takeout": "horse_win_takeout",
    "Horse Win Hold": "horse_win_hold",
    "Horse Place Takeout": "horse_place_takeout",
    "Horse PlaceHold": "horse_place_hold",
    "Race Hold": "race_hold",
    "BetBack Information": "betback_information",
    "BetBack Flag": "betback_flag",
    "Refund Flag": "refund_flag",
    "Placing": "placing_position",
    "Win Deduction": "win_deduction",
    "Place Deduction": "place_deduction",
    "Win Result": "win_result",
    "Place Result": "place_result",
    "Win Payout Amount": "win_payout_amount",
    "Place Payout Amount": "place_payout_amount",
    "Paid Status": "paid_status",
    "Bet Terminal": "bet_terminal",
}

TABLE_COLUMNS = [
    "source_file",
    "source_sheet",
    "source_row_number",
    "refid",
    "source_uuid",
    "license",
    "bookmaker",
    "location",
    "state",
    "area",
    "wagering_provider",
    "sys_ver_no",
    "bet_date",
    "bet_time",
    "event_date",
    "event_time",
    "ticket_no",
    "event_category",
    "sport_event",
    "venue",
    "venue_state",
    "bet_type",
    "bet_method",
    "bet_details",
    "race_number",
    "runner_number",
    "runner_name",
    "bet_amount_win",
    "bet_amount_place",
    "win_price",
    "place_price",
    "customer_name",
    "betback_claim",
    "bet_information",
    "cancelled_flag",
    "time_cancelled",
    "bet_win_takeout",
    "bet_place_takeout",
    "horse_win_takeout",
    "horse_win_hold",
    "horse_place_takeout",
    "horse_place_hold",
    "race_hold",
    "betback_information",
    "betback_flag",
    "refund_flag",
    "placing_position",
    "win_deduction",
    "place_deduction",
    "win_result",
    "place_result",
    "win_payout_amount",
    "place_payout_amount",
    "paid_status",
    "bet_terminal",
    "raw_payload",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Load bookmaker Excel dumps into raw.bookmaker_bet_dump."
    )
    parser.add_argument(
        "--raw-dir",
        type=Path,
        default=RAW_DATA_DIR,
        help="Directory containing source .xlsx files.",
    )
    parser.add_argument(
        "--apply-schema",
        action="store_true",
        help="Apply the raw schema before loading data.",
    )
    parser.add_argument(
        "--replace-files",
        action="store_true",
        help="Delete existing rows for each source file before reloading it.",
    )
    parser.add_argument(
        "--limit-files",
        type=int,
        default=None,
        help="Only load the first N workbook files for debugging.",
    )
    parser.add_argument(
        "--limit-rows",
        type=int,
        default=None,
        help="Only load the first N data rows from each workbook for debugging.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Validate and preview workbook ingestion without connecting to PostgreSQL.",
    )
    return parser.parse_args()


def stringify_cell(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    return str(value).strip()


def normalize_headers(headers: Iterable[object]) -> list[str]:
    return [str(header).strip() for header in headers]


def validate_headers(headers: list[str]) -> None:
    unknown_headers = [header for header in headers if header not in HEADER_TO_COLUMN]
    if unknown_headers:
        raise ValueError(f"Unknown source headers found: {unknown_headers}")


def empty_record() -> dict[str, str | None]:
    return {column: None for column in TABLE_COLUMNS if column not in {"source_file", "source_sheet", "source_row_number", "raw_payload"}}


def build_record(headers: list[str], values: tuple[object, ...]) -> tuple[dict[str, str | None], dict[str, str | None]]:
    record = empty_record()
    raw_payload: dict[str, str | None] = {}

    for header, value in zip(headers, values):
        text_value = stringify_cell(value)
        raw_payload[header] = text_value
        record[HEADER_TO_COLUMN[header]] = text_value

    return record, raw_payload


def iter_workbook_rows(path: Path, limit_rows: int | None = None) -> Iterator[tuple[str, int, dict[str, str | None], dict[str, str | None]]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)

    headers = normalize_headers(next(rows))
    validate_headers(headers)

    for row_number, values in enumerate(rows, start=2):
        if values is None or not any(value is not None for value in values):
            continue

        record, raw_payload = build_record(headers, values)
        yield sheet.title, row_number, record, raw_payload

        if limit_rows is not None and (row_number - 1) >= limit_rows:
            break


def preview_record(record: dict[str, str | None]) -> str:
    preview_fields = {
        "refid": record["refid"],
        "bookmaker": record["bookmaker"],
        "bet_date": record["bet_date"],
        "bet_time": record["bet_time"],
        "sport_event": record["sport_event"],
        "bet_type": record["bet_type"],
        "bet_amount_win": record["bet_amount_win"],
        "customer_name": record["customer_name"],
    }
    return json.dumps(preview_fields, ensure_ascii=True)


def dry_run_workbooks(args: argparse.Namespace) -> None:
    workbook_paths = sorted(args.raw_dir.glob("*.xlsx"))
    if args.limit_files is not None:
        workbook_paths = workbook_paths[: args.limit_files]

    if not workbook_paths:
        raise FileNotFoundError(f"No .xlsx files found in {args.raw_dir}")

    total_rows = 0

    for workbook_path in workbook_paths:
        row_count = 0
        sheet_name: str | None = None
        sample_preview: str | None = None

        for source_sheet, _, record, _ in iter_workbook_rows(
            workbook_path,
            limit_rows=args.limit_rows,
        ):
            sheet_name = source_sheet
            row_count += 1
            if sample_preview is None:
                sample_preview = preview_record(record)

        total_rows += row_count
        print(
            f"[DRY RUN] {workbook_path.name} | sheet={sheet_name or 'N/A'} | rows={row_count}"
        )
        if sample_preview is not None:
            print(f"  sample={sample_preview}")

    print(f"[DRY RUN] Validated {len(workbook_paths)} workbook(s), {total_rows} row(s) total.")


def chunked(items: list[tuple], batch_size: int = 500) -> Iterator[list[tuple]]:
    for start in range(0, len(items), batch_size):
        yield items[start : start + batch_size]


def record_to_tuple(
    source_file: str,
    source_sheet: str,
    source_row_number: int,
    record: dict[str, str | None],
    raw_payload: dict[str, str | None],
    json_wrapper,
) -> tuple:
    return (
        source_file,
        source_sheet,
        source_row_number,
        record["refid"],
        record["source_uuid"],
        record["license"],
        record["bookmaker"],
        record["location"],
        record["state"],
        record["area"],
        record["wagering_provider"],
        record["sys_ver_no"],
        record["bet_date"],
        record["bet_time"],
        record["event_date"],
        record["event_time"],
        record["ticket_no"],
        record["event_category"],
        record["sport_event"],
        record["venue"],
        record["venue_state"],
        record["bet_type"],
        record["bet_method"],
        record["bet_details"],
        record["race_number"],
        record["runner_number"],
        record["runner_name"],
        record["bet_amount_win"],
        record["bet_amount_place"],
        record["win_price"],
        record["place_price"],
        record["customer_name"],
        record["betback_claim"],
        record["bet_information"],
        record["cancelled_flag"],
        record["time_cancelled"],
        record["bet_win_takeout"],
        record["bet_place_takeout"],
        record["horse_win_takeout"],
        record["horse_win_hold"],
        record["horse_place_takeout"],
        record["horse_place_hold"],
        record["race_hold"],
        record["betback_information"],
        record["betback_flag"],
        record["refund_flag"],
        record["placing_position"],
        record["win_deduction"],
        record["place_deduction"],
        record["win_result"],
        record["place_result"],
        record["win_payout_amount"],
        record["place_payout_amount"],
        record["paid_status"],
        record["bet_terminal"],
        json_wrapper(raw_payload),
    )


def apply_schema(cursor, schema_path: Path) -> None:
    cursor.execute(schema_path.read_text(encoding="utf-8"))


def load_workbooks(args: argparse.Namespace) -> None:
    import psycopg
    from psycopg.types.json import Json

    workbook_paths = sorted(args.raw_dir.glob("*.xlsx"))
    if args.limit_files is not None:
        workbook_paths = workbook_paths[: args.limit_files]

    if not workbook_paths:
        raise FileNotFoundError(f"No .xlsx files found in {args.raw_dir}")

    db_config = get_database_config()
    insert_sql = f"""
        insert into raw.bookmaker_bet_dump ({", ".join(TABLE_COLUMNS)})
        values ({", ".join(["%s"] * len(TABLE_COLUMNS))})
    """

    total_inserted = 0

    with psycopg.connect(
        host=db_config.host,
        port=db_config.port,
        dbname=db_config.dbname,
        user=db_config.user,
        password=db_config.password,
    ) as conn:
        with conn.cursor() as cursor:
            if args.apply_schema:
                apply_schema(cursor, SCHEMA_PATH)

            for workbook_path in workbook_paths:
                if args.replace_files:
                    cursor.execute(
                        "delete from raw.bookmaker_bet_dump where source_file = %s",
                        (workbook_path.name,),
                    )

                batch: list[tuple] = []
                inserted_for_file = 0

                for source_sheet, source_row_number, record, raw_payload in iter_workbook_rows(
                    workbook_path,
                    limit_rows=args.limit_rows,
                ):
                    batch.append(
                        record_to_tuple(
                            source_file=workbook_path.name,
                            source_sheet=source_sheet,
                            source_row_number=source_row_number,
                            record=record,
                            raw_payload=raw_payload,
                            json_wrapper=Json,
                        )
                    )

                    if len(batch) >= 500:
                        cursor.executemany(insert_sql, batch)
                        inserted_for_file += len(batch)
                        total_inserted += len(batch)
                        batch.clear()

                if batch:
                    cursor.executemany(insert_sql, batch)
                    inserted_for_file += len(batch)
                    total_inserted += len(batch)

                print(f"Loaded {inserted_for_file} rows from {workbook_path.name}")

        conn.commit()

    print(f"Finished loading {total_inserted} rows into raw.bookmaker_bet_dump")


def main() -> None:
    args = parse_args()
    if args.dry_run:
        dry_run_workbooks(args)
        return
    load_workbooks(args)


if __name__ == "__main__":
    main()
